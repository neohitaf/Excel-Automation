import openpyxl
from openpyxl.styles import Font
from templates import BORDRO_SABLON_YOLU, BANKA_LISTESI_SABLON_YOLU, MUHTASAR_SABLON_YOLU, MUHTASAR_SABIT
from datetime import datetime
import calendar

def gss_ayir(kayitlar):
    evet_list = []
    hayir_list = []
    for k in kayitlar:
        if k.get("gss") == "EVET":   
            evet_list.append(k)
        else:
            hayir_list.append(k)
    return evet_list, hayir_list


def banka_listesi_yaz(kayitlar, ay, cikti_yolu):
    
    

    AY_NUMARALARI = {"OCAK": 1, "ŞUBAT": 2, "MART": 3, "NİSAN": 4,
                     "MAYIS": 5, "HAZİRAN": 6, "TEMMUZ": 7, "AĞUSTOS": 8,
                     "EYLÜL": 9, "EKİM": 10, "KASIM": 11, "ARALIK": 12}
    yil = datetime.now().year
    ay_no = AY_NUMARALARI.get(ay, 1)
    son_gun = calendar.monthrange(yil, ay_no)[1]
    donem = f"01 {ay}- {son_gun} {ay} {yil}"

    evet_list, hayir_list = gss_ayir(kayitlar)
    sirali = evet_list + hayir_list

    wb = openpyxl.load_workbook(BANKA_LISTESI_SABLON_YOLU)
    ws = wb.active

    # Dönem satırını güncelle
    ws.cell(row=5, column=4, value=donem)

    # Veri satırları — 9. satırdan başla
    satir_no = 9
    for k in sirali:
        ws.cell(row=satir_no, column=2, value=k.get("ad_soyad", ""))
        ws.cell(row=satir_no, column=3, value=k.get("tc", ""))
        ws.cell(row=satir_no, column=4, value=k.get("birim", ""))
        ws.cell(row=satir_no, column=5, value=k.get("iban", ""))
        satir_no += 1

    wb.save(cikti_yolu)
    print(f"Kaydedildi: {cikti_yolu}")

def bordro_yaz(kayitlar, ay, cikti_yolu):
    AY_NUMARALARI = {"OCAK": 1, "ŞUBAT": 2, "MART": 3, "NİSAN": 4,
                     "MAYIS": 5, "HAZİRAN": 6, "TEMMUZ": 7, "AĞUSTOS": 8,
                     "EYLÜL": 9, "EKİM": 10, "KASIM": 11, "ARALIK": 12}
    yil = datetime.now().year
    ay_no = AY_NUMARALARI.get(ay, 1)
    son_gun = calendar.monthrange(yil, ay_no)[1]
    baslik_tarihi = f"1 {ay} {yil} - {son_gun} {ay} {yil}"

    evet_list, hayir_list = gss_ayir(kayitlar)

    wb = openpyxl.load_workbook(BORDRO_SABLON_YOLU)
    ws = wb.active

    ws.cell(row=2, column=1, value=baslik_tarihi)

    satir_no = 6
    for k in evet_list:
        tc = k.get("tc", "") or "eksik"
        ws.cell(row=satir_no, column=2, value=k.get("ad_soyad", ""))
        ws.cell(row=satir_no, column=3, value=tc)
        ws.cell(row=satir_no, column=4, value=k.get("birim", ""))
        ws.cell(row=satir_no, column=5, value=1101)
        ws.cell(row=satir_no, column=6, value=k.get("prim_gunu", 0))
        satir_no += 1

    satir_no = 84
    for k in hayir_list:
        tc = k.get("tc", "") or "eksik"
        ws.cell(row=satir_no, column=2, value=k.get("ad_soyad", ""))
        ws.cell(row=satir_no, column=3, value=tc)
        ws.cell(row=satir_no, column=4, value=k.get("birim", ""))
        ws.cell(row=satir_no, column=5, value=1101)
        ws.cell(row=satir_no, column=6, value=k.get("prim_gunu", 0))
        satir_no += 1

    wb.save(cikti_yolu)
    print(f"Kaydedildi: {cikti_yolu}")


def muhtasar_yaz(kayitlar, ay, cikti_yolu):


    AY_NUMARALARI = {"OCAK": 1, "ŞUBAT": 2, "MART": 3, "NİSAN": 4,
                     "MAYIS": 5, "HAZİRAN": 6, "TEMMUZ": 7, "AĞUSTOS": 8,
                     "EYLÜL": 9, "EKİM": 10, "KASIM": 11, "ARALIK": 12}
    yil = datetime.now().year
    ay_no = AY_NUMARALARI.get(ay, 1)
    ay_metin = f"{ay_no:02d}"

    evet_list, hayir_list = gss_ayir(kayitlar)
    sirali = evet_list + hayir_list

    wb = openpyxl.load_workbook(MUHTASAR_SABLON_YOLU)
    ws = wb["MUHTASAR "]

    satir_no = 2
    for k in sirali:
        tc = k.get("tc", "")

        ws.cell(row=satir_no, column=1,  value=MUHTASAR_SABIT["belgenin_mahiyeti"])
        ws.cell(row=satir_no, column=2,  value=k.get("belge_turu", ""))
        ws.cell(row=satir_no, column=3,  value=MUHTASAR_SABIT["kanun_no"])
        ws.cell(row=satir_no, column=4,  value=MUHTASAR_SABIT["yeni_unite_kodu"])
        ws.cell(row=satir_no, column=5,  value=MUHTASAR_SABIT["eski_unite_kodu"])
        ws.cell(row=satir_no, column=6,  value=MUHTASAR_SABIT["isyeri_sira_no"])
        ws.cell(row=satir_no, column=7,  value=MUHTASAR_SABIT["il_kodu"])
        ws.cell(row=satir_no, column=8,  value=MUHTASAR_SABIT["alt_isveren_no"])
        ws.cell(row=satir_no, column=9,  value=tc)
        ws.cell(row=satir_no, column=10, value=tc)
        ws.cell(row=satir_no, column=11, value=k.get("ad", ""))
        ws.cell(row=satir_no, column=12, value=k.get("soyad", ""))
        ws.cell(row=satir_no, column=13, value=k.get("prim_gunu", 0))
        ws.cell(row=satir_no, column=15, value=k.get("ucret", 0))
        ws.cell(row=satir_no, column=22, value=k.get("eksik_gun", 0))          # V
        ws.cell(row=satir_no, column=23, value=MUHTASAR_SABIT["eksik_gun_nedeni"])
        ws.cell(row=satir_no, column=24, value=MUHTASAR_SABIT["meslek_kodu"])
        ws.cell(row=satir_no, column=26, value=MUHTASAR_SABIT["tahakkuk_nedeni"])
        ws.cell(row=satir_no, column=27, value=ay_metin)
        ws.cell(row=satir_no, column=28, value=str(yil))
        ws.cell(row=satir_no, column=29, value=MUHTASAR_SABIT["gelir_vergisi_muaf"])
        ws.cell(row=satir_no, column=31, value=MUHTASAR_SABIT["gv_engellilik_orani"])  # AE
        ws.cell(row=satir_no, column=34, value=MUHTASAR_SABIT["gv_kesintisi"])         # AH
        ws.cell(row=satir_no, column=36, value=MUHTASAR_SABIT["damga_vergisi_kesintisi"])  # AJ
        # AD, AF, AG, AI'ye hiç dokunmuyoruz — formülleri korunuyor

        satir_no += 1

    wb.save(cikti_yolu)
    print(f"Kaydedildi: {cikti_yolu}")