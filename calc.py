import openpyxl
from parser import parse_tablo
from templates import FORM3_TEMPLATE
from writer import banka_listesi_yaz, bordro_yaz

GUNLUK_UCRET = 1101
SGK_AYI = 30


def puantaj_oku(dosya_yolu):
    """
    Puantaj Excel'ini okur.cle
    """

    wb = openpyxl.load_workbook(dosya_yolu, data_only=True)
    ws = wb.active

    # Birim ve ay ara
    birim = ""
    ay = ""
    for row in ws.iter_rows(max_row=15):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if "Birimin Adı" in val or "Kurum Adı" in val:
                birim = val.replace("Birimin Adı:", "").replace("Kurum Adı:", "").strip()
            if val in ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
                       "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]:
                ay = val

    # Başlık satırını bul — "Adı Soyadı" veya "Ad Soyad" geçen satır
    baslik_satiri = None
    sutun_haritasi = {}
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            val = str(cell.value).strip().upper() if cell.value else ""
            if "ADI SOYADI" in val or "AD SOYAD" in val or "SOYAD" in val:
                baslik_satiri = cell.row
                break
        if baslik_satiri:
            break

    if not baslik_satiri:
        return [], ay

    # Başlık satırındaki sütun isimlerini haritala
    for cell in ws[baslik_satiri]:
        val = str(cell.value).strip().upper() if cell.value else ""
        if "ADI SOYADI" in val or "AD SOYAD" in val:
            sutun_haritasi["ad_soyad"] = cell.column - 1
        elif "GÖREV" in val or "GÖREVİ" in val:
            sutun_haritasi["gorev"] = cell.column - 1
        elif "TOPLAM" in val:
            sutun_haritasi["toplam"] = cell.column - 1

    kayitlar = []
    for row in ws.iter_rows(min_row=baslik_satiri + 1):
        ad_soyad = row[sutun_haritasi["ad_soyad"]].value if "ad_soyad" in sutun_haritasi else None
        gorev    = row[sutun_haritasi["gorev"]].value    if "gorev"    in sutun_haritasi else None
        toplam   = row[sutun_haritasi["toplam"]].value   if "toplam"   in sutun_haritasi else None

        if not ad_soyad:
            continue
        if not gorev:
            continue
        if str(gorev).strip().lower() != "öğrenci":
            continue

        kayitlar.append({
            "ad_soyad":    str(ad_soyad).strip(),
            "toplam_saat": int(toplam) if toplam else 0,
            "birim":       birim,
        })

    return kayitlar, ay
 

def turkce_upper(s):
    return s.replace("i", "İ").replace("ı", "I").upper()


def esle(parser_record, puantaj_record):
    """
    Puantaj kaydı ile parser kaydını eşleştirir.
    Döner: (eşleşen parser kaydı, uyarı)
    """
    ad_soyad = puantaj_record["ad_soyad"]
    for name in parser_record:
        if name["ad_soyad"].upper() == ad_soyad.upper():
            return name, None
    return None, f"'{ad_soyad}' Form-3'te bulunamadı."
    
    


def hesapla(kayit):
    saat        = kayit.get("toplam_saat", 0)   
    prim_gunu   = saat // 8
    ucret       = prim_gunu * GUNLUK_UCRET
    eksik_gun   = SGK_AYI - prim_gunu
    belge_turu  = 22 if kayit.get("gss") == "EVET" else 43

    return {
        **kayit,
        "prim_gunu":   prim_gunu,
        "ucret":       ucret,
        "eksik_gun":   eksik_gun,
        "belge_turu":  belge_turu,
    }

if __name__ == "__main__":
    parser_kayitlari, uyarilar = parse_tablo("veri/Form-3.docx", FORM3_TEMPLATE)
    kayitlar, ay = puantaj_oku("veri/Puantaj_ornek.xlsx")

    # BANKA LİSTESİ — sadece Form-3'ten üret, puantajla kıyaslama yok
    banka_sonuclari = []
    for parser_kaydi in parser_kayitlari:
        sonuc = hesapla(parser_kaydi)
        banka_sonuclari.append(sonuc)

    # BORDRO — Puantaj merkezli, eksik varsa bildir
    bordro_sonuclari = []
    for puantaj_kaydi in kayitlar:
        parser_kaydi, uyari = esle(parser_kayitlari, puantaj_kaydi)
        if uyari:
            print("UYARI (Bordro):", uyari)
            parser_kaydi = {}
        birlesik = {**parser_kaydi, **puantaj_kaydi}
        sonuc = hesapla(birlesik)
        bordro_sonuclari.append(sonuc)

    banka_listesi_yaz(banka_sonuclari, ay, "cikti/banka_listesi.xlsx")
    bordro_yaz(bordro_sonuclari, ay, "cikti/bordro.xlsx")